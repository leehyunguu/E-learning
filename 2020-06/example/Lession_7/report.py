# -*- coding: utf-8 -*-

import openpyxl
# Excel 파일에서 셀의 스타일 설정을 변경하기 위한 라이브러리들을 등록합니다.
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color

# Excel 파일에서 차트의 스타일 설정을 변경하기 위한 라이브러리들을 등록합니다.
from openpyxl.chart import BarChart, Reference, Series

def set_title(ws):
    # sheet 에서 특정 셀을 선택해서 병합 정렬을 수행합니다.
    ws.merge_cells('A1:E1')

    # sheet 에서 A1 셀에 'Report' 라는 값을 입력합니다.
    ws['A1'] = 'Report'
    title = ws['A1']

    # sheet 의 font 설정을 Font 모듈을 사용해서 설정합니다.
    title.font = Font(name='nanum gothic', size=18, bold=True)

    # sheet 의 정렬 설정을 Alignment 모듈을 사용해서 설정합니다.
    title.alignment = Alignment(horizontal='center', vertical='center')

    # sheet 의 fill 설정을 PatternFill 모듈을 사용해서 설정합니다.
    # 셀의 뒷 배경의 색상을 변경할 수 있습니다.
    title.fill = PatternFill(patternType='solid', fgColor=Color('808080'))

    # Border 모듈을 사용해서, 셀의 테두리 설정을 할 수 있습니다.
    # 셀의 왼쪽, 오른쪽, 윗쪽, 아래쪽에 대한 설정을 개별적으로 할 수 있습니다.
    box = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # 설정한 테두리 값을 border 설정에 적용합니다.
    ws['A1'].border = box
    ws['B1'].border = box
    ws['C1'].border = box
    ws['D1'].border = box
    ws['E1'].border = box

def set_data(ws):
    # Excel 에 표시할 기본 데이터 값을 설정합니다.
    # 여기서는 rows 로 설정을 하고, 각각의 rows 를 순차적으로 더하는 방식을 사용합니다.
    rows = [
        ['V1', 'V2', 'V3', 'V4', 'V5'],
        [20, 90, 15, 80, 75],
        [30, 80, 25, 45, 80],
        [40, 70, 55, 88, 75],
        [50, 60, 75, 87, 89],
        [60, 50, 35, 85, 55],
        [70, 40, 5, 76, 85],
    ]

    # 사전에 설정한 rows 값을 sheet 를 돌면서 값을 추가합니다.
    for row in rows:
        ws.append(row)

def set_chart(ws):
    # 차트의 레이블에 사용할 데이터의 위치를 설정합니다.
    # - ws : Excel sheet 가 저장된 변수
    # - min_col : Excel sheet 에서 데이터를 읽기 시작하는 위치의 행
    # - min_row : Excel sheet 에서 데이터를 읽기 시작하는 위치의 열
    #   - Title 과 데이터의 레이블을 제외하여, 1+2 = 3 이 됨
    # - max_row : Excel sheet 에서 데이터가 있는 마지막 부분의 행
    categories = Reference(ws, min_col=1, min_row=3, max_row=8)

    # 차트에서 보여줄 데이터의 위치를 설정합니다.
    # - ws : Excel sheet 가 저장된 변수
    # - min_col : Excel sheet 에서 데이터를 읽기 시작하는 위치의 행
    # - min_row : Excel sheet 에서 데이터를 읽기 시작하는 위치의 열
    #   - Title 을 제외하고, 데이터의 레이블을 포함해서 2 가 됨
    # - max_col : Excel sheet 에서 데이터가 있는 마지막 부분의 행
    # - max_row : Excel sheet 에서 데이터가 있는 마지막 부분의 열
    values = Reference(ws, min_col=1, min_row=2, max_col=5, max_row=8)

    # BarChart 를 생성합니다.
    chart = BarChart()

    # add_data 를 사용해서 앞서 설정한 values 를 넘겨줍니다.
    # data 에 제목이 포함되어있으면, titles_from_data 를 True 로 세팅합니다.
    chart.add_data(values, titles_from_data=True)

    # set_categories 를 사용해서 앞서 설정한 categories 를 설정합니다.
    chart.set_categories(categories)

    # 차트가 생성될 위치를 지정합니다.
    ws.add_chart(chart, 'A10')


def main():
    # openpyxl 의 Workbook 모듈을 사용해서, 새로운 Excel 데이터를 만듭니다.
    wb = openpyxl.Workbook()

    # workbook 에서 activate 메소드를 사용하면, 현재 활성화되어있는 sheet 를 가져옵니다.
    # 기본 값으로 sheet1 이 선택됩니다.
    ws = wb.active

    # title 메소드를 사용하면, sheet 의 이름을 설정할 수 있습니다.
    ws.title = 'Report'

    set_title(ws)
    set_data(ws)
    set_chart(ws)

    # 설정한 sheet 가 저장된 excel 을 저장합니다.
    wb.save('report.xlsx')
    wb.close()


if __name__ == '__main__':
    main()
