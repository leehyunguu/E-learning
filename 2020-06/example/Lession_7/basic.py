# -*- coding: utf-8 -*-

import openpyxl

def main():
    # openpyxl 의 load_workbook 메소드를 사용해서 excel 파일을 불러옵니다.
    # openpyxl 은 엑셀이 설치되어 있지 않아도, excel 파일을 읽을 수 있습니다.
    wb = openpyxl.load_workbook('sample.xlsx')

    # load_workbook 으로 불러온 excel 파일에서 sheet 이름을 가지고 각각의 sheet 를 조작할 수 있습니다.
    ws = wb['Sheet1']

    # 불러온 sheet 에서 excel 의 각 행들을 rows 메소드를 사용해서 순차적으로 접근할 수 있습니다.
    for r in ws.rows:
        # 아래 부분은 excel 의 첫번째 행에 제목을 작성하기 위한 부분입니다.
        # 다양한 방법으로 행에 접근할 수 있다는 것을 보여주기 위해 아래와 같은 방법을 사용했습니다.
        # 간단하게는 다음과 같이 확인할 수도 있습니다.
        #* - r[1].value
        # - 이와 같은 방법은 조회만 가능하고, 값을 새로 쓰는 것은 할 수 없습니다.
        # - 예를 들어, r[5].value = 'Sum' 와 같이 사용하면, IndexError 오류가 발생합니다.
        row_index = r[0].row
        if row_index == 1:
            ws.cell(row=row_index, column=5).value = 'Sum'
            ws.cell(row=row_index, column=6).value = 'Avg'
            continue

        math = r[1].value
        eng = r[2].value
        history = r[3].value
        summary = math + eng + history

        ws.cell(row=row_index, column=5).value = summary
        ws.cell(row=row_index, column=6).value = summary / 3
        # cell 의 숫자 속성을 '#.#' 으로 변경합니다.
        ws.cell(row=row_index, column=6).number_format = '#.#'

    # 변경한 workbook(excel) 이 내용을 저장합니다.
    wb.save('sample2.xlsx')
    wb.close()

if __name__ == '__main__':
    main()
